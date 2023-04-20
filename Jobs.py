import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date

class Jobs:

    my_jobs_tags = []
    job_list = []

    @classmethod
    def add_job_tag(self, tag):
        self.my_jobs_tags.append(tag)

    @classmethod
    def get_all_tags(self):
        for idx, tag in enumerate(self.my_jobs_tags):
            print(f"{idx+1}. {tag}")
        input("press enter to continue")

    @classmethod
    def get_jobs_from_pracujpl(self):
        for tag in self.my_jobs_tags:
            job_dict = {}
            tag = tag.replace(" ", "%20")
            url = f"https://www.pracuj.pl/praca/{tag};kw/krakow;wp?rd=50&et=1%2C17"
            page = requests.get(url)
            soup = BeautifulSoup(page.content, 'html.parser')

            for job_data in soup.find_all('div', class_="c8i823f"):
                job_dict = {}
                job_dict["job_name"] = job_data.find('h2', class_="b1iadbg8").text
                job_dict["company_name"] = job_data.find('h4', class_="e1ml1ys4 t1c1o3wg").text
                href_index_start = str(job_data.find('h2', class_="b1iadbg8")).find(" href=\"https:")
                href_index_end = str(job_data.find('h2', class_="b1iadbg8")).find("\">", href_index_start)
                offer_url = str(job_data.find('h2', class_="b1iadbg8"))[href_index_start + 7:href_index_end]
                if """https://""" not in offer_url:
                    offer_url = "error, propobly offer splited to locations"
                job_dict["offer_url"] = offer_url
                job_dict["source"] = "pracuj.pl"
                today = date.today()
                job_dict["get_date"] = today.strftime("%d/%m/%Y")
                job_dict["tag_name"] = tag
                self.job_list.append(job_dict)
        input("press enter to continue")

    @classmethod
    def export_to_xml(self):
        try:
            wb = load_workbook('job_db.xlsx')
            ws = wb["Pracujpl"]
            print("updated old exel file")
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = "Pracujpl"
            ws['B1'] = 'Job title'
            ws['C1'] = 'Company'
            ws['D1'] = 'Found for tag'
            ws['E1'] = 'Finding date'
            ws['F1'] = 'Finding site'
            ws['G1'] = 'Offer_url'
            ws['H1'] = 'Applicated'
            ws['I1'] = 'Not qualified to apply'


            for row_nr, data in enumerate(self.job_list):
                ws.append(
                    ["", data["job_name"], data["company_name"], data["tag_name"], data["get_date"], data["source"],
                     data["offer_url"]])
            print("created new exel_file")
        wb.save('job_db.xlsx')

    @classmethod
    def backup_to_txt(self,backup_name):
        with open("backups/"+backup_name+".txt", "w") as file:
            file.write(str(self.my_jobs_tags))

    @classmethod
    def backup_from_txt(self,backup_name):
        with open("backups/" + backup_name + ".txt", "r") as file:
            file = file.read()
            self.my_jobs_tags = file.strip('][').replace('\'', "").split(', ')

